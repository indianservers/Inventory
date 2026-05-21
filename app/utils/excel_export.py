from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def export_to_excel(headers: list, rows: list[list], sheet_name: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Report")[:31]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_index, row in enumerate(rows, start=2):
        ws.append(row)
        if row_index % 2 == 0:
            for cell in ws[row_index]:
                cell.fill = alt_fill

    ws.freeze_panes = "A2"
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(width, 12), 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

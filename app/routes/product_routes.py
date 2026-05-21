import base64
import csv
from pathlib import Path
from io import BytesIO, StringIO

import qrcode
from flask import Blueprint, Response, current_app, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import Brand, Category, Product, ProductImage, ProductVariant, Supplier, Tax, Unit, Warehouse
from app.utils.tax_validation import is_valid_hsn

bp = Blueprint("products", __name__, url_prefix="/products")


def load_choices():
    return {
        "categories": Category.query.order_by(Category.name).all(),
        "brands": Brand.query.order_by(Brand.name).all(),
        "units": Unit.query.order_by(Unit.name).all(),
        "taxes": Tax.query.order_by(Tax.name).all(),
        "warehouses": Warehouse.query.order_by(Warehouse.name).all(),
        "suppliers": Supplier.query.order_by(Supplier.name).all(),
    }


def save_product(product):
    sku = request.form["sku"].strip()
    existing = Product.query.filter(Product.sku == sku, Product.id != (product.id or 0)).first()
    if existing:
        raise ValueError("SKU already exists.")
    product.sku = sku
    product.barcode = request.form.get("barcode") or _barcode_from_sku(sku)
    product.name = request.form["name"]
    product.description = request.form.get("description")
    product.category_id = request.form.get("category_id") or None
    product.brand_id = request.form.get("brand_id") or None
    product.unit_id = request.form.get("unit_id") or None
    product.tax_id = request.form.get("tax_id") or None
    product.preferred_supplier_id = request.form.get("preferred_supplier_id") or None
    if not is_valid_hsn(request.form.get("hsn_code")):
        raise ValueError("Invalid HSN/SAC code. Use 4 to 8 digits.")
    product.hsn_code = request.form.get("hsn_code")
    for field in ["purchase_price", "sales_price", "mrp", "opening_stock", "min_stock", "max_stock", "reorder_level"]:
        setattr(product, field, request.form.get(field) or 0)
    if not product.id:
        product.current_stock = product.opening_stock
        product.average_cost = product.purchase_price
        product.created_by = current_user.id
    product.warehouse_id = request.form.get("warehouse_id") or None
    product.rack_bin = request.form.get("rack_bin")
    product.track_inventory = bool(request.form.get("track_inventory"))
    product.batch_tracking = bool(request.form.get("batch_tracking"))
    product.serial_tracking = bool(request.form.get("serial_tracking"))
    product.expiry_tracking = bool(request.form.get("expiry_tracking"))
    product.is_active = bool(request.form.get("is_active"))


def _barcode_from_sku(sku):
    digits = "".join(str(ord(ch)) for ch in sku)[-10:].rjust(10, "0")
    return f"89{digits}"


def save_variants(product):
    existing = {str(variant.id): variant for variant in product.variants.all()} if product.id else {}
    seen = set()
    for variant_id, sku, barcode, size, color, weight, model, packaging, purchase_price, sales_price, mrp, current_stock in zip(
        request.form.getlist("variant_id[]"),
        request.form.getlist("variant_sku[]"),
        request.form.getlist("variant_barcode[]"),
        request.form.getlist("variant_size[]"),
        request.form.getlist("variant_color[]"),
        request.form.getlist("variant_weight[]"),
        request.form.getlist("variant_model[]"),
        request.form.getlist("variant_packaging[]"),
        request.form.getlist("variant_purchase_price[]"),
        request.form.getlist("variant_sales_price[]"),
        request.form.getlist("variant_mrp[]"),
        request.form.getlist("variant_stock[]"),
    ):
        sku = (sku or "").strip()
        if not sku:
            continue
        duplicate = ProductVariant.query.filter(ProductVariant.sku == sku, ProductVariant.id != int(variant_id or 0)).first()
        if duplicate:
            raise ValueError(f"Variant SKU already exists: {sku}")
        variant = existing.get(variant_id) or ProductVariant(product_id=product.id)
        variant.sku = sku
        variant.barcode = barcode or _barcode_from_sku(sku)
        variant.size = size
        variant.color = color
        variant.weight = weight
        variant.model = model
        variant.packaging = packaging
        variant.purchase_price = purchase_price or 0
        variant.sales_price = sales_price or 0
        variant.mrp = mrp or 0
        variant.current_stock = current_stock or 0
        variant.is_active = True
        db.session.add(variant)
        if variant_id:
            seen.add(variant_id)
    for variant_id, variant in existing.items():
        if variant_id not in seen:
            variant.is_active = False


def save_product_images(product):
    files = [file for file in request.files.getlist("images") if file and file.filename]
    if not files:
        return
    upload_dir = Path(current_app.config["UPLOAD_FOLDER"]) / "products"
    upload_dir.mkdir(parents=True, exist_ok=True)
    has_primary = product.images.filter_by(is_primary=True).first() is not None
    for file in files:
        extension = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if extension not in {"png", "jpg", "jpeg", "gif", "webp"}:
            continue
        filename = secure_filename(f"{product.sku}-{product.id}-{len(files)}-{file.filename}")
        target = upload_dir / filename
        file.save(target)
        db.session.add(ProductImage(product_id=product.id, image_path=f"uploads/products/{filename}", is_primary=not has_primary))
        has_primary = True


@bp.route("/")
@login_required
def index():
    search = request.args.get("q", "").strip()
    category_id = request.args.get("category_id")
    stock_status = request.args.get("stock_status", "all")
    query = Product.query
    if search:
        like = f"%{search}%"
        query = query.filter(
            db.or_(
                Product.sku.ilike(like),
                Product.barcode.ilike(like),
                Product.name.ilike(like),
                Product.hsn_code.ilike(like),
                Product.rack_bin.ilike(like),
            )
        )
    if category_id:
        query = query.filter(Product.category_id == category_id)
    if stock_status == "low":
        query = query.filter(Product.min_stock > 0, Product.current_stock <= Product.min_stock)
    elif stock_status == "out":
        query = query.filter(Product.current_stock <= 0)
    elif stock_status == "active":
        query = query.filter(Product.is_active.is_(True))
    elif stock_status == "inactive":
        query = query.filter(Product.is_active.is_(False))
    products = query.order_by(Product.id.desc()).all()
    selected_product = Product.query.get(request.args.get("selected")) if request.args.get("selected") else (products[0] if products else None)
    return render_template("products/index.html", title="Item Master", products=products, selected_product=selected_product, search=search, categories=Category.query.order_by(Category.name).all(), category_id=category_id, stock_status=stock_status)


@bp.route("/export")
@login_required
def export():
    fmt = request.args.get("format", "csv")
    products = Product.query.order_by(Product.sku).all()
    headers = [
        "sku", "barcode", "name", "description", "hsn_code", "purchase_price",
        "sales_price", "mrp", "current_stock", "average_cost", "opening_stock",
        "min_stock", "max_stock", "reorder_level", "rack_bin", "is_active",
    ]
    rows = [[getattr(product, field) for field in headers] for product in products]
    if fmt == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=products.xlsx"},
        )

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=products.csv"},
    )


@bp.route("/import", methods=["GET", "POST"])
@login_required
def import_products():
    if request.method == "POST":
        upload = request.files.get("file")
        if not upload or not upload.filename:
            flash("Choose a CSV or Excel file.", "danger")
            return redirect(url_for("products.import_products"))

        created = updated = 0
        rows = _read_product_rows(upload)
        for row in rows:
            sku = (row.get("sku") or "").strip()
            name = (row.get("name") or "").strip()
            if not sku or not name:
                continue
            product = Product.query.filter_by(sku=sku).first()
            if product:
                updated += 1
            else:
                product = Product(sku=sku, created_by=current_user.id)
                db.session.add(product)
                created += 1
            _apply_import_row(product, row)
        db.session.commit()
        flash(f"Import complete. Created {created}, updated {updated}.", "success")
        return redirect(url_for("products.index"))
    return render_template("products/import.html", title="Import Products")


def _read_product_rows(upload):
    filename = upload.filename.lower()
    if filename.endswith(".xlsx"):
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]

    stream = StringIO(upload.stream.read().decode("utf-8-sig"))
    return list(csv.DictReader(stream))


def _apply_import_row(product, row):
    product.barcode = row.get("barcode") or product.barcode
    product.name = row.get("name") or product.name
    product.description = row.get("description") or product.description
    product.hsn_code = row.get("hsn_code") or product.hsn_code
    product.rack_bin = row.get("rack_bin") or product.rack_bin
    for field in ["purchase_price", "sales_price", "mrp", "opening_stock", "min_stock", "max_stock", "reorder_level"]:
        if row.get(field) not in (None, ""):
            setattr(product, field, row.get(field))
    if row.get("current_stock") not in (None, ""):
        product.current_stock = row.get("current_stock")
    if row.get("average_cost") not in (None, ""):
        product.average_cost = row.get("average_cost")
    if row.get("is_active") not in (None, ""):
        product.is_active = str(row.get("is_active")).lower() in {"1", "true", "yes", "active"}


@bp.route("/<int:id>/qr")
@login_required
def qr_label(id):
    product = Product.query.get_or_404(id)
    payload = f"{product.sku}|{product.barcode or ''}|{product.name}"
    image = qrcode.make(payload)
    output = BytesIO()
    image.save(output, format="PNG")
    qr_data = base64.b64encode(output.getvalue()).decode("ascii")
    return render_template("products/qr.html", title=f"QR Label - {product.sku}", product=product, qr_data=qr_data)


@bp.route("/create", methods=["GET", "POST"])
@login_required
def create():
    product = Product()
    if request.method == "POST":
        try:
            save_product(product)
            db.session.add(product)
            db.session.flush()
            save_variants(product)
            save_product_images(product)
            db.session.commit()
            flash("Product created.", "success")
            return redirect(url_for("products.index"))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template("products/form.html", title="Create Product", product=product, **load_choices())


@bp.route("/<int:id>/edit", methods=["GET", "POST"])
@login_required
def edit(id):
    product = Product.query.get_or_404(id)
    if request.method == "POST":
        try:
            save_product(product)
            save_variants(product)
            save_product_images(product)
            db.session.commit()
            flash("Product updated.", "success")
            return redirect(url_for("products.index"))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template("products/form.html", title="Edit Product", product=product, **load_choices())


@bp.route("/images/<int:id>/primary")
@login_required
def set_primary_image(id):
    image = ProductImage.query.get_or_404(id)
    ProductImage.query.filter_by(product_id=image.product_id).update({"is_primary": False})
    image.is_primary = True
    db.session.commit()
    flash("Primary image updated.", "success")
    return redirect(url_for("products.edit", id=image.product_id))


@bp.route("/images/<int:id>/delete")
@login_required
def delete_image(id):
    image = ProductImage.query.get_or_404(id)
    product_id = image.product_id
    db.session.delete(image)
    db.session.commit()
    flash("Product image deleted.", "success")
    return redirect(url_for("products.edit", id=product_id))


@bp.route("/<int:id>/delete")
@login_required
def delete(id):
    product = Product.query.get_or_404(id)
    product.is_active = False
    db.session.commit()
    flash("Product marked inactive.", "success")
    return redirect(url_for("products.index"))


@bp.route("/barcode-labels")
@login_required
def barcode_labels():
    ids = request.args.getlist("id")
    query = Product.query
    if ids:
        query = query.filter(Product.id.in_(ids))
    products = query.order_by(Product.name).limit(100).all()
    labels = []
    for product in products:
        payload = product.barcode or product.sku
        image = qrcode.make(payload)
        output = BytesIO()
        image.save(output, format="PNG")
        labels.append({"product": product, "barcode": payload, "qr_data": base64.b64encode(output.getvalue()).decode("ascii")})
    return render_template("products/barcode_labels.html", title="Barcode Labels", products=Product.query.order_by(Product.name).all(), labels=labels)
